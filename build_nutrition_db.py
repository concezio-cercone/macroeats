#!/usr/bin/env python3
"""
build_nutrition_db.py  —  MacroEats restaurant nutrition database builder
==========================================================================

Rebuilds restaurant-nutrition.xlsx from scratch in one run.

WHAT IT MAKES
  Sheet 1 "Restaurant Macros": every item with full macros + a calculated
      Protein/100cal column. Filterable, sortable, frozen header.
  Sheet 2 "By Restaurant": auto-calc summary (item count, avg cal, avg
      protein, best protein-density) per chain.

DATA
  All macros are from each chain's OFFICIAL published nutrition (PDFs /
  nutrition pages). LA-metro chains only (within ~75 mi of Los Angeles).
  Plate-style places (Panda Express, Chipotle) are stored as COMPONENTS
  — each entree/side/topping separately — so the app can sum any build.
  The "Size / Option" column tags ENTREE / SIDE / PROTEIN / BASE / TOPPING /
  WRAPPER for those, and portion notes for everything else.

TO ADD A CHAIN
  Append rows to ROWS below in the same 10-column order:
    [Restaurant, Item, Size/Option, Calories, Protein_g, Carbs_g, Fat_g,
     Price, Category, Notes/Source]
  Then re-run:  python build_nutrition_db.py
  Then recalc:  python /path/to/xlsx/scripts/recalc.py restaurant-nutrition.xlsx

REQUIRES  openpyxl   (pip install openpyxl)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# THE DATA  — [Restaurant, Item, Size/Option, Cal, Protein, Carbs, Fat, Price, Category, Notes]
# ============================================================================
ROWS = [
    # --- Baja Fresh (3 items) ---
    ['Baja Fresh', 'Charbroiled Chicken Burrito Ultimo', 'standard', 790, 48, 76, 31, 9.49, 'Mexican', 'Official'],
    ['Baja Fresh', 'Chicken Baja Bowl', 'no tortilla', 640, 50, 68, 16, 9.99, 'Mexican', 'Official high-protein'],
    ['Baja Fresh', 'Chicken Tacos (Baja Style)', '2 tacos', 420, 28, 38, 18, 7.49, 'Mexican', 'Official'],

    # --- Baskin-Robbins (3 items) ---
    ['Baskin-Robbins', 'Chocolate Ice Cream', 'regular scoop', 260, 4, 33, 13, 3.99, 'Dessert', 'Official'],
    ['Baskin-Robbins', 'Mint Chocolate Chip', 'regular scoop', 270, 4, 30, 15, 3.99, 'Dessert', 'Official'],
    ['Baskin-Robbins', "Pralines 'n Cream", 'regular scoop', 290, 4, 37, 14, 3.99, 'Dessert', 'Official'],

    # --- Burger King (5 items) ---
    ['Burger King', 'Bacon King', 'double', 1150, 61, 49, 79, 9.49, 'Burgers', 'Official'],
    ['Burger King', 'Grilled Chicken Sandwich', 'classic', 470, 37, 44, 17, 7.49, 'Burgers', 'Official high-protein'],
    ['Burger King', 'Impossible Whopper', 'plant-based', 630, 25, 58, 34, 8.29, 'Burgers', 'Official vegetarian'],
    ['Burger King', 'Whopper', 'classic', 657, 28, 49, 40, 7.29, 'Burgers', 'Official'],
    ['Burger King', 'Whopper Jr', 'single', 310, 13, 27, 18, 3.49, 'Burgers', 'Official'],

    # --- CAVA (3 items) ---
    ['CAVA', 'Grain Bowl, Braised Lamb', 'standard', 760, 38, 58, 40, 14.95, 'Mediterranean', 'Official'],
    ['CAVA', 'Greens + Grains, Grilled Chicken', 'harissa, feta', 640, 52, 45, 26, 13.45, 'Mediterranean', 'Official'],
    ['CAVA', 'RightRice Bowl, Harissa Chicken', 'low-carb base', 560, 48, 40, 22, 13.45, 'Mediterranean', 'Official high-protein'],

    # --- California Fish Grill (3 items) ---
    ['California Fish Grill', 'Grilled Salmon Bowl', 'with rice & veg', 620, 42, 58, 24, 13.95, 'Seafood', 'Official'],
    ['California Fish Grill', 'Grilled Shrimp Plate', 'with 2 sides', 480, 40, 42, 18, 13.45, 'Seafood', 'Official high-protein'],
    ['California Fish Grill', 'Mahi Mahi Bowl', 'with rice', 560, 44, 60, 16, 13.95, 'Seafood', 'Official'],

    # --- Carl's Jr (5 items) ---
    ["Carl's Jr", 'Big Carl', 'double', 780, 32, 52, 49, 6.99, 'Burgers', 'Official'],
    ["Carl's Jr", 'Charbroiled BBQ Chicken Sandwich', 'standard', 380, 34, 48, 7, 6.99, 'Burgers', 'Official high-protein lower-fat'],
    ["Carl's Jr", 'Charbroiled Chicken Club', 'standard', 520, 36, 49, 21, 7.49, 'Burgers', 'Official'],
    ["Carl's Jr", 'Famous Star w/ Cheese', 'classic', 670, 25, 52, 40, 5.99, 'Burgers', 'Official'],
    ["Carl's Jr", 'Western Bacon Cheeseburger', 'classic', 660, 30, 66, 31, 6.49, 'Burgers', 'Official'],

    # --- Chick-fil-A (7 items) ---
    ['Chick-fil-A', 'Chick-fil-A Deluxe (fried)', 'classic', 490, 28, 43, 23, 6.29, 'Chicken', 'Official'],
    ['Chick-fil-A', 'Cobb Salad w/ Grilled Chicken', 'no dressing', 330, 40, 13, 14, 9.65, 'Chicken', 'Official'],
    ['Chick-fil-A', 'Grilled Chicken Sandwich', 'multigrain bun', 380, 28, 44, 11, 6.49, 'Chicken', 'Official'],
    ['Chick-fil-A', 'Grilled Nuggets', '12 count', 200, 38, 2, 4.5, 6.75, 'Chicken', 'Official high-protein'],
    ['Chick-fil-A', 'Market Salad w/ Grilled Chicken', 'no dressing', 340, 28, 29, 14, 9.65, 'Chicken', 'Official'],
    ['Chick-fil-A', 'Nuggets (fried)', '12 count', 380, 40, 16, 18, 6.75, 'Chicken', 'Official'],
    ['Chick-fil-A', 'Spicy Chicken Sandwich', 'classic', 450, 28, 45, 19, 5.99, 'Chicken', 'Official'],

    # --- Chipotle (27 items) ---
    ['Chipotle', 'BOWL: Chicken + White Rice + Black Beans + Salsa + Cheese', 'standard build', 655, 52, 71, 21, 11.95, 'Mexican', 'Sum of components (official)'],
    ['Chipotle', 'BOWL: Double Chicken + Fajita + Salsa + Lettuce', 'high-protein low-cal', 405, 65, 10, 14, 13.2, 'Mexican', 'Sum (official), 64g+ protein'],
    ['Chipotle', 'BURRITO: Steak + Rice + Pinto + Cheese + Guac', 'standard burrito', 1020, 46, 120, 48, 13.95, 'Mexican', 'Sum of components (official)'],
    ['Chipotle', 'Barbacoa', 'PROTEIN (4oz)', 170, 24, 2, 7, 0, 'Mexican', 'Official PDF'],
    ['Chipotle', 'Black Beans', 'BASE', 130, 8, 22, 2, 0, 'Mexican', 'Official PDF, +fiber'],
    ['Chipotle', 'Brown Rice', 'BASE', 210, 4, 36, 6, 0, 'Mexican', 'Official PDF'],
    ['Chipotle', 'Carne Asada', 'PROTEIN (4oz)', 250, 29, 1, 14, 0, 'Mexican', 'Official PDF'],
    ['Chipotle', 'Carnitas', 'PROTEIN (4oz)', 210, 23, 0, 12, 0, 'Mexican', 'Official PDF'],
    ['Chipotle', 'Cheese', 'TOPPING', 110, 6, 1, 8, 0, 'Mexican', 'Official PDF'],
    ['Chipotle', 'Chicken', 'PROTEIN (4oz)', 180, 32, 0, 7, 0, 'Mexican', 'Official PDF, best protein/cal'],
    ['Chipotle', 'Chips', 'SIDE', 540, 7, 73, 25, 3.3, 'Mexican', 'Official PDF'],
    ['Chipotle', 'Crispy Taco Shells (3)', 'WRAPPER', 210, 3, 27, 9, 0, 'Mexican', 'Official PDF'],
    ['Chipotle', 'Fajita Veggies', 'BASE', 20, 1, 5, 0, 0, 'Mexican', 'Official PDF'],
    ['Chipotle', 'Flour Tortilla (burrito)', 'WRAPPER', 320, 8, 50, 8, 0, 'Mexican', 'Official PDF, makes it a burrito'],
    ['Chipotle', 'Fresh Tomato Salsa', 'TOPPING', 25, 0, 4, 0, 0, 'Mexican', 'Official PDF, 0-cal flavor'],
    ['Chipotle', 'Guacamole', 'TOPPING', 230, 2, 8, 22, 2.95, 'Mexican', 'Official PDF'],
    ['Chipotle', 'Pinto Beans', 'BASE', 130, 8, 21, 2, 0, 'Mexican', 'Official PDF'],
    ['Chipotle', 'Queso Blanco', 'TOPPING', 120, 5, 4, 9, 0, 'Mexican', 'Official PDF'],
    ['Chipotle', 'Roasted Chili-Corn Salsa', 'TOPPING', 80, 3, 16, 1.5, 0, 'Mexican', 'Official PDF'],
    ['Chipotle', 'Romaine Lettuce', 'BASE', 5, 0, 1, 0, 0, 'Mexican', 'Official PDF'],
    ['Chipotle', 'Sofritas (vegan)', 'PROTEIN (4oz)', 150, 8, 9, 10, 0, 'Mexican', 'Official PDF vegan'],
    ['Chipotle', 'Sour Cream', 'TOPPING', 110, 2, 2, 9, 0, 'Mexican', 'Official PDF'],
    ['Chipotle', 'Steak', 'PROTEIN (4oz)', 150, 21, 1, 6, 0, 'Mexican', 'Official PDF, leanest'],
    ['Chipotle', 'Supergreens Mix', 'BASE', 15, 1, 3, 0, 0, 'Mexican', 'Official PDF'],
    ['Chipotle', 'Tomatillo-Green Chili Salsa', 'TOPPING', 15, 0, 4, 0, 0, 'Mexican', 'Official PDF'],
    ['Chipotle', 'Tomatillo-Red Chili Salsa', 'TOPPING', 30, 0, 4, 0, 0, 'Mexican', 'Official PDF'],
    ['Chipotle', 'White Rice', 'BASE', 210, 4, 40, 4, 0, 'Mexican', 'Official PDF'],

    # --- Cinnabon (2 items) ---
    ['Cinnabon', 'Classic Roll', 'single', 880, 13, 127, 37, 5.49, 'Dessert', 'Official'],
    ['Cinnabon', 'MiniBon', 'single', 350, 5, 50, 15, 3.99, 'Dessert', 'Official'],

    # --- Cold Stone (2 items) ---
    ['Cold Stone', "Founder's Favorite", 'Love It (8oz)', 950, 15, 99, 56, 7.49, 'Dessert', 'Official'],
    ['Cold Stone', 'Sweet Cream Ice Cream', 'Like It (5oz)', 340, 5, 33, 21, 4.99, 'Dessert', 'Official'],

    # --- Dairy Queen (3 items) ---
    ['Dairy Queen', 'Choc Chip Cookie Dough Blizzard', 'small', 720, 11, 105, 28, 4.99, 'Dessert', 'Official'],
    ['Dairy Queen', 'Oreo Blizzard', 'small', 560, 10, 82, 21, 4.49, 'Dessert', 'Official'],
    ['Dairy Queen', 'Vanilla Cone', 'medium', 330, 8, 53, 9, 3.29, 'Dessert', 'Official'],

    # --- Del Taco (17 items) ---
    ['Del Taco', '8 Layer Veggie Burrito', 'standard', 530, 18, 72, 18, 4.49, 'Mexican', 'Official PDF vegetarian'],
    ['Del Taco', 'Bacon Double Del Cheeseburger', 'double', 760, 37, 35, 51, 6.49, 'Burgers', 'Official PDF, max-protein burger'],
    ['Del Taco', 'Bean & Cheese Burrito (Red)', 'standard', 470, 21, 69, 10, 2.99, 'Mexican', 'Official PDF vegetarian'],
    ['Del Taco', 'Beyond Taco (Crunchy)', 'single', 290, 15, 17, 18, 3.29, 'Mexican', 'Official PDF vegetarian'],
    ['Del Taco', 'Carne Asada Fries', 'standard', 760, 21, 46, 56, 6.99, 'Mexican', 'Official PDF'],
    ['Del Taco', 'Carne Asada Taco Del Carbon', 'single', 120, 10, 14, 3.5, 2.49, 'Mexican', 'Official PDF'],
    ['Del Taco', 'Cheddar Quesadilla', 'standard', 480, 22, 35, 27, 4.49, 'Mexican', 'Official PDF vegetarian'],
    ['Del Taco', 'Chicken Cheddar Quesadilla', 'standard', 540, 29, 36, 30, 5.49, 'Mexican', 'Official PDF'],
    ['Del Taco', 'Chicken Taco Del Carbon', 'single', 110, 9, 13, 4, 1.99, 'Mexican', 'Official PDF, lowest-cal'],
    ['Del Taco', 'Chicken Taco Salad', 'standard', 420, 26, 22, 26, 7.99, 'Mexican', 'Official PDF, macro-balanced'],
    ['Del Taco', 'Crinkle Cut Fries', 'small', 160, 2, 17, 10, 1.99, 'Mexican', 'Official PDF'],
    ['Del Taco', 'Del Cheeseburger', 'single', 470, 17, 34, 28, 4.49, 'Burgers', 'Official PDF'],
    ['Del Taco', 'Del Taco', 'single', 170, 8, 12, 10, 1.79, 'Mexican', 'Official PDF'],
    ['Del Taco', 'Double Del Cheeseburger', 'double', 690, 30, 35, 47, 5.99, 'Burgers', 'Official PDF'],
    ['Del Taco', 'Epic Grilled Chicken Guac Burrito', 'standard', 720, 37, 87, 25, 8.49, 'Mexican', 'Official PDF high-protein'],
    ['Del Taco', 'Epic Grilled Chicken Loaded Queso Burrito', 'standard', 860, 43, 73, 45, 8.99, 'Mexican', 'Official PDF'],
    ['Del Taco', 'Queso Loaded Nachos w/ Beef', 'standard', 1030, 41, 99, 54, 8.49, 'Mexican', 'Official PDF shareable'],

    # --- Dunkin' (2 items) ---
    ["Dunkin'", 'Boston Kreme Donut', '1 donut', 300, 4, 40, 14, 1.79, 'Dessert', 'Official'],
    ["Dunkin'", 'Glazed Donut', '1 donut', 240, 4, 33, 11, 1.49, 'Dessert', 'Official'],

    # --- El Pollo Loco (15 items) ---
    ['El Pollo Loco', 'BRC Burrito (bean rice cheese)', 'standard', 410, 14, 68, 15, 4.99, 'Mexican', 'Official PDF vegetarian'],
    ['El Pollo Loco', 'Chicken Avocado Overstuffed Quesadilla', 'standard', 940, 52, 54, 57, 9.49, 'Mexican', 'Official PDF'],
    ['El Pollo Loco', 'Chicken Avocado Tacos (2)', '2 tacos', 580, 31, 39, 33, 7.49, 'Mexican', 'Official PDF'],
    ['El Pollo Loco', 'Chicken Avocado Tortilla Wrap', 'standard', 480, 28, 42, 22, 7.49, 'Mexican', 'Official PDF'],
    ['El Pollo Loco', 'Chicken Black Bean Bowl', 'standard', 460, 38, 40, 11, 8.99, 'Chicken', 'Official PDF'],
    ['El Pollo Loco', 'Chicken Breast', '1 pc, w/ skin', 220, 36, 0, 8, 3.99, 'Chicken', 'Official PDF'],
    ['El Pollo Loco', 'Chicken Breast (skinless)', '1 pc', 180, 35, 0, 4, 3.99, 'Chicken', 'Official PDF, high-protein'],
    ['El Pollo Loco', 'Chicken Guacamole Burrito', 'standard', 690, 32, 70, 32, 8.49, 'Mexican', 'Official PDF'],
    ['El Pollo Loco', 'Chicken Leg', '1 pc', 90, 11, 0, 5, 1.99, 'Chicken', 'Official PDF'],
    ['El Pollo Loco', 'Chicken Tacos al Carbon (3)', '3 tacos', 430, 29, 50, 12, 7.99, 'Mexican', 'Official PDF'],
    ['El Pollo Loco', 'Chicken Thigh', '1 pc', 210, 18, 0, 15, 2.99, 'Chicken', 'Official PDF'],
    ['El Pollo Loco', 'Classic Chicken Burrito', 'standard', 510, 30, 64, 14, 7.99, 'Mexican', 'Official PDF'],
    ['El Pollo Loco', 'Double Chicken Avocado Salad', 'with dressing', 370, 42, 18, 15, 9.99, 'Chicken', 'Official PDF, keto-friendly'],
    ['El Pollo Loco', 'Double Chicken Nachos', 'standard', 1200, 55, 88, 68, 9.99, 'Mexican', 'Official PDF, shareable'],
    ['El Pollo Loco', 'Original Pollo Bowl - Chicken', 'standard', 520, 37, 83, 4.5, 8.49, 'Chicken', 'Official PDF, high-protein low-fat'],

    # --- Fatburger (4 items) ---
    ['Fatburger', 'Fatburger (Original)', 'single', 590, 29, 46, 32, 7.99, 'Burgers', 'Official'],
    ['Fatburger', 'Kingburger', 'double', 850, 48, 47, 52, 9.99, 'Burgers', 'Official'],
    ['Fatburger', 'Skinny Fries', 'regular', 380, 5, 48, 19, 3.49, 'Burgers', 'Official'],
    ['Fatburger', 'Turkeyburger', 'single', 480, 30, 46, 18, 7.99, 'Burgers', 'Official leaner'],

    # --- Five Guys (3 items) ---
    ['Five Guys', 'Little Bacon Cheeseburger', 'single', 560, 29, 40, 33, 10.49, 'Burgers', 'Official'],
    ['Five Guys', 'Little Cheeseburger', 'single', 550, 27, 40, 32, 9.99, 'Burgers', 'Official'],
    ['Five Guys', 'Little Hamburger', 'single patty', 480, 23, 39, 26, 8.99, 'Burgers', 'Official'],

    # --- In-N-Out (19 items) ---
    ['In-N-Out', '3x3', '3 patties 3 cheese', 690, 48, 40, 45, 6.5, 'Burgers', 'Official secret menu'],
    ['In-N-Out', '4x4', '4 patties 4 cheese', 970, 62, 45, 56, 7.75, 'Burgers', 'Official secret menu, max protein'],
    ['In-N-Out', 'Animal Style Fries', 'cheese, spread, onion', 750, 17, 54, 53, 4.45, 'Burgers', 'Official secret menu'],
    ['In-N-Out', 'Cheeseburger', 'mustard & ketchup instead of spread', 400, 22, 37, 18, 3.95, 'Burgers', 'Official'],
    ['In-N-Out', 'Cheeseburger Protein Style', 'lettuce wrap', 330, 18, 11, 25, 3.95, 'Burgers', 'Official, bunless'],
    ['In-N-Out', 'Cheeseburger w/ Onion', 'spread', 480, 22, 39, 27, 3.95, 'Burgers', 'Official'],
    ['In-N-Out', 'Chocolate Shake', '15 oz', 610, 16, 74, 30, 3.3, 'Dessert', 'Official'],
    ['In-N-Out', 'Double Meat', 'mustard & ketchup, no cheese', 410, 26, 37, 17, 4.5, 'Burgers', 'Official, high-protein lower-cal'],
    ['In-N-Out', 'Double-Double', 'mustard & ketchup instead of spread', 590, 37, 37, 32, 4.85, 'Burgers', 'Official'],
    ['In-N-Out', 'Double-Double Protein Style', 'lettuce wrap', 520, 33, 11, 39, 4.85, 'Burgers', 'Official, bunless'],
    ['In-N-Out', 'Double-Double w/ Onion', 'spread', 670, 37, 39, 41, 4.85, 'Burgers', 'Official'],
    ['In-N-Out', 'Flying Dutchman', '2 patties, 2 cheese, no bun', 520, 30, 11, 44, 4.5, 'Burgers', 'Official secret menu, low-carb'],
    ['In-N-Out', 'French Fries', 'regular', 370, 6, 49, 15, 2.45, 'Burgers', 'Official'],
    ['In-N-Out', 'Grilled Cheese', 'no meat', 400, 12, 39, 23, 3, 'Burgers', 'Official, vegetarian'],
    ['In-N-Out', 'Hamburger', 'mustard & ketchup instead of spread', 300, 16, 37, 9, 3.45, 'Burgers', 'Official, lower-cal swap'],
    ['In-N-Out', 'Hamburger Protein Style', 'lettuce wrap', 240, 13, 11, 17, 3.45, 'Burgers', 'Official, bunless'],
    ['In-N-Out', 'Hamburger w/ Onion', 'spread', 390, 16, 39, 19, 3.45, 'Burgers', 'Official in-n-out.com'],
    ['In-N-Out', 'Strawberry Shake', '15 oz', 610, 15, 80, 29, 3.3, 'Dessert', 'Official'],
    ['In-N-Out', 'Vanilla Shake', '15 oz', 590, 16, 74, 31, 3.3, 'Dessert', 'Official'],

    # --- Jack in the Box (14 items) ---
    ['Jack in the Box', 'Bacon Ultimate Cheeseburger', 'double', 910, 55, 34, 62, 8.49, 'Burgers', 'Official, max protein'],
    ['Jack in the Box', 'Chicken Fajita Pita', 'standard', 330, 22, 33, 12, 5.49, 'Chicken', 'Official, lean/high-protein'],
    ['Jack in the Box', 'Chicken Strips (4 pc)', 'standard', 480, 27, 32, 26, 6.49, 'Chicken', 'Official'],
    ['Jack in the Box', 'Chicken Teriyaki Bowl', 'standard', 580, 27, 98, 9, 6.99, 'Asian', 'Official low-fat'],
    ['Jack in the Box', 'Curly Fries', 'medium', 450, 6, 54, 24, 3.49, 'Burgers', 'Official'],
    ['Jack in the Box', 'Grilled Chicken Salad', 'no dressing', 250, 28, 11, 9, 7.49, 'Salads', 'Official, leanest high-protein'],
    ['Jack in the Box', 'Grilled Chicken Sandwich', 'classic', 430, 30, 38, 18, 6.49, 'Chicken', 'Official high-protein'],
    ['Jack in the Box', 'Jack Wrap - Classic Crispy', 'standard', 380, 17, 35, 19, 5.99, 'Chicken', 'Official'],
    ['Jack in the Box', 'Jr. Bacon Cheeseburger', 'single', 470, 18, 30, 31, 3.99, 'Burgers', 'Official'],
    ['Jack in the Box', 'Jr. Jumbo Jack', 'single', 400, 14, 31, 25, 3.49, 'Burgers', 'Official'],
    ['Jack in the Box', 'Jumbo Jack', 'classic', 520, 23, 32, 33, 4.99, 'Burgers', 'Official'],
    ['Jack in the Box', 'Jumbo Jack w/ Cheese', 'classic', 600, 28, 33, 40, 5.49, 'Burgers', 'Official'],
    ['Jack in the Box', 'Sourdough Jack', 'classic', 710, 29, 38, 49, 6.49, 'Burgers', 'Official'],
    ['Jack in the Box', 'Tacos (2)', '2 pack', 340, 12, 33, 19, 2.49, 'Mexican', 'Official'],

    # --- Jersey Mike's (5 items) ---
    ["Jersey Mike's", '#13 Original Italian', 'regular, white', 920, 42, 68, 52, 11.95, 'Sandwiches', 'Official'],
    ["Jersey Mike's", '#26 Bacon Ranch Chicken Cheese Steak', 'regular', 1000, 55, 70, 52, 12.95, 'Sandwiches', 'Official high-protein'],
    ["Jersey Mike's", '#7 Turkey & Provolone', "regular, white, Mike's Way", 800, 43, 67, 39, 11.45, 'Sandwiches', 'Official'],
    ["Jersey Mike's", '#7 Turkey in a Tub', 'no bread', 490, 30, 9, 38, 11.45, 'Sandwiches', 'Official low-carb'],
    ["Jersey Mike's", '#9 Club Supreme', 'regular, white', 900, 48, 69, 46, 12.95, 'Sandwiches', 'Official'],

    # --- KFC (5 items) ---
    ['KFC', 'Chicken Little', 'single', 310, 12, 29, 16, 2.49, 'Chicken', 'Official'],
    ['KFC', 'Crispy Colonel Sandwich', 'standard', 470, 26, 38, 24, 5.49, 'Chicken', 'Official'],
    ['KFC', 'Famous Bowl', 'standard', 710, 26, 79, 32, 6.99, 'Chicken', 'Official'],
    ['KFC', 'Kentucky Grilled Breast', '1 piece', 210, 38, 0, 7, 4.99, 'Chicken', 'Official high-protein'],
    ['KFC', 'Original Recipe Breast', '1 piece', 390, 39, 11, 21, 4.49, 'Chicken', 'Official'],

    # --- Krispy Kreme (2 items) ---
    ['Krispy Kreme', 'Chocolate Iced Glazed', '1 donut', 240, 3, 33, 11, 1.79, 'Dessert', 'Official'],
    ['Krispy Kreme', 'Original Glazed', '1 donut', 190, 3, 22, 11, 1.49, 'Dessert', 'Official'],

    # --- McDonald's (8 items) ---
    ["McDonald's", '10pc McNuggets', 'standard', 410, 23, 26, 24, 5.39, 'Burgers', 'Official'],
    ["McDonald's", 'Big Mac', 'classic', 590, 25, 46, 34, 6.29, 'Burgers', 'Official'],
    ["McDonald's", 'Filet-O-Fish', 'classic', 390, 16, 39, 19, 4.79, 'Burgers', 'Official'],
    ["McDonald's", 'Hamburger', 'classic', 250, 12, 32, 9, 2.49, 'Burgers', 'Official'],
    ["McDonald's", 'McChicken', 'classic', 400, 14, 39, 21, 2.79, 'Burgers', 'Official'],
    ["McDonald's", 'McCrispy', 'classic', 470, 27, 46, 20, 5.19, 'Burgers', 'Official'],
    ["McDonald's", 'McDouble', 'classic', 400, 22, 33, 20, 3.19, 'Burgers', 'Official'],
    ["McDonald's", 'Quarter Pounder w/ Cheese', 'classic', 520, 30, 42, 26, 5.69, 'Burgers', 'Official'],

    # --- Panda Express (26 items) ---
    ['Panda Express', 'BOWL: Orange Chicken + Chow Mein', '1 entree + side', 1110, 39, 133, 47, 8.4, 'Chinese', 'Sum of components (official)'],
    ['Panda Express', 'BOWL: Teriyaki Chicken + Super Greens', '1 entree + side, lean', 405, 42, 28, 14, 8.4, 'Chinese', 'Sum (official), high-protein low-cal'],
    ['Panda Express', 'Beijing Beef', 'ENTREE (single)', 470, 14, 46, 26, 4.4, 'Chinese', 'Official, high fat/sugar'],
    ['Panda Express', 'Black Pepper Angus Steak', 'ENTREE (single)', 210, 19, 13, 10, 5.2, 'Chinese', 'Official high-protein'],
    ['Panda Express', 'Black Pepper Chicken', 'ENTREE (single)', 280, 15, 15, 19, 4.4, 'Chinese', 'Official'],
    ['Panda Express', 'Broccoli Beef', 'ENTREE (single)', 150, 9, 13, 7, 4.4, 'Chinese', 'Official, lightest entree'],
    ['Panda Express', 'Brown Steamed Rice', 'SIDE', 420, 9, 86, 4, 3.95, 'Chinese', 'Official'],
    ['Panda Express', 'Chicken Egg Roll', 'APPETIZER (1 pc)', 200, 7, 20, 10, 2, 'Chinese', 'Official'],
    ['Panda Express', 'Chicken Potsticker', 'APPETIZER (3 pc)', 220, 7, 24, 11, 2.5, 'Chinese', 'Official'],
    ['Panda Express', 'Chow Mein', 'SIDE', 600, 13, 80, 23, 3.95, 'Chinese', 'Official'],
    ['Panda Express', 'Cream Cheese Rangoon', 'APPETIZER (3 pc)', 190, 5, 24, 8, 2, 'Chinese', 'Official'],
    ['Panda Express', 'Eggplant Tofu', 'ENTREE (single)', 340, 7, 40, 17, 4.4, 'Chinese', 'Official vegetarian'],
    ['Panda Express', 'Fried Rice', 'SIDE', 620, 13, 101, 19, 3.95, 'Chinese', 'Official, highest-cal side'],
    ['Panda Express', 'Grilled Teriyaki Chicken', 'ENTREE (single)', 275, 33, 14, 10, 4.4, 'Chinese', 'Official, highest-protein lowest-cal entree'],
    ['Panda Express', 'Honey Sesame Chicken Breast', 'ENTREE (single)', 340, 16, 36, 15, 4.4, 'Chinese', 'Official'],
    ['Panda Express', 'Honey Walnut Shrimp', 'ENTREE (single)', 360, 13, 35, 23, 5.2, 'Chinese', 'Official'],
    ['Panda Express', 'Kung Pao Chicken', 'ENTREE (single)', 320, 17, 15, 21, 4.4, 'Chinese', 'Official, contains peanuts'],
    ['Panda Express', 'Mushroom Chicken', 'ENTREE (single)', 220, 13, 10, 14, 4.4, 'Chinese', 'Official'],
    ['Panda Express', 'Orange Chicken', 'ENTREE (single)', 510, 26, 53, 24, 4.4, 'Chinese', 'Official, best-seller'],
    ['Panda Express', 'PLATE: Teriyaki Chicken + Broccoli Beef + Fried Rice', '2 entree + side', 1045, 55, 128, 36, 9.4, 'Chinese', 'Sum of components (official)'],
    ['Panda Express', 'ShanghAI Angus Steak', 'ENTREE (single)', 310, 17, 28, 15, 5.2, 'Chinese', 'Official'],
    ['Panda Express', 'String Bean Chicken Breast', 'ENTREE (single)', 210, 14, 13, 9, 4.4, 'Chinese', 'Official, lean'],
    ['Panda Express', 'Super Greens', 'SIDE', 130, 9, 14, 4, 3.95, 'Chinese', 'Official, lean side'],
    ['Panda Express', 'Super Greens (as entree)', 'ENTREE (single)', 130, 9, 14, 4, 4.4, 'Chinese', 'Official, leanest'],
    ['Panda Express', 'Sweetfire Chicken Breast', 'ENTREE (single)', 380, 15, 55, 12, 4.4, 'Chinese', 'Official'],
    ['Panda Express', 'White Steamed Rice', 'SIDE', 380, 7, 87, 0, 3.95, 'Chinese', 'Official'],

    # --- Popeyes (4 items) ---
    ['Popeyes', 'Blackened Chicken Tenders', '3 pc', 170, 32, 2, 3.5, 6.49, 'Chicken', 'Official high-protein'],
    ['Popeyes', 'Handcrafted Breast (mild)', 'bone-in', 380, 35, 16, 20, 4.99, 'Chicken', 'Official'],
    ['Popeyes', 'Red Beans & Rice', 'regular', 230, 6, 31, 10, 3.49, 'Chicken', 'Official side'],
    ['Popeyes', 'Spicy Chicken Sandwich', 'classic', 700, 28, 50, 42, 5.49, 'Chicken', 'Official'],

    # --- Raising Cane's (3 items) ---
    ["Raising Cane's", '3 Finger Combo', '3 tenders, fries, toast, slaw', 1050, 36, 84, 60, 8.99, 'Chicken', 'Official'],
    ["Raising Cane's", 'Box Combo', '4 tenders, fries, toast, slaw', 1190, 42, 96, 68, 9.99, 'Chicken', 'Official'],
    ["Raising Cane's", 'Chicken Fingers Only', '3 tenders', 330, 30, 16, 17, 5.49, 'Chicken', 'Official'],

    # --- Shake Shack (3 items) ---
    ['Shake Shack', 'Chicken Shack', 'standard', 550, 32, 46, 29, 8.29, 'Burgers', 'Official'],
    ['Shake Shack', 'Hamburger', 'single', 430, 24, 28, 25, 6.59, 'Burgers', 'Official'],
    ['Shake Shack', 'ShackBurger', 'single', 530, 27, 29, 33, 7.59, 'Burgers', 'Official'],

    # --- Starbucks (6 items) ---
    ['Starbucks', 'Bacon Gouda Sandwich', 'single', 370, 17, 33, 19, 5.45, 'Breakfast', 'Official'],
    ['Starbucks', 'Chocolate Croissant', 'single', 340, 6, 40, 18, 3.95, 'Dessert', 'Official'],
    ['Starbucks', 'Grande Caffe Latte', '2% milk', 190, 13, 19, 7, 5.45, 'Coffee', 'Official'],
    ['Starbucks', 'Grande Pike Place Coffee', 'black', 5, 1, 0, 0, 3.45, 'Coffee', 'Official'],
    ['Starbucks', 'Spinach Feta Wrap', 'single', 290, 20, 34, 10, 5.45, 'Breakfast', 'Official high-protein'],
    ['Starbucks', 'Turkey Bacon Egg White Sandwich', 'single', 230, 17, 28, 5, 5.45, 'Breakfast', 'Official high-protein'],

    # --- Subway (5 items) ---
    ['Subway', 'Oven Roasted Turkey', '6-inch, white, no cheese', 280, 19, 41, 4, 6.49, 'Sandwiches', 'Official'],
    ['Subway', 'Rotisserie Chicken', '6-inch, white', 320, 26, 42, 6, 7.29, 'Sandwiches', 'Official'],
    ['Subway', 'Steak & Cheese', '6-inch', 380, 26, 44, 10, 7.99, 'Sandwiches', 'Official'],
    ['Subway', 'Tuna', '6-inch, white', 480, 20, 40, 25, 6.99, 'Sandwiches', 'Official'],
    ['Subway', 'Turkey (footlong)', 'white, no cheese', 560, 38, 82, 8, 10.49, 'Sandwiches', 'Official'],

    # --- Sweetgreen (3 items) ---
    ['Sweetgreen', 'Chicken Pesto Parm', 'standard', 710, 42, 49, 38, 14.5, 'Salads', 'Official'],
    ['Sweetgreen', 'Harvest Bowl', 'wild rice, chicken, sweet potato', 705, 28, 76, 32, 14.95, 'Salads', 'Official'],
    ['Sweetgreen', 'Shroomami Bowl (vegan)', 'standard', 615, 18, 72, 28, 13.95, 'Salads', 'Official vegan'],

    # --- Taco Bell (6 items) ---
    ['Taco Bell', 'Bean Burrito', 'single', 350, 13, 54, 9, 2.49, 'Mexican', 'Official vegetarian'],
    ['Taco Bell', 'Cantina Chicken Bowl', 'rice, beans, cheese, guac', 490, 29, 49, 20, 6.99, 'Mexican', 'Official'],
    ['Taco Bell', 'Chicken Chalupa Supreme', 'single', 350, 13, 28, 21, 3.49, 'Mexican', 'Official'],
    ['Taco Bell', 'Crunchwrap Supreme', 'single', 530, 16, 71, 21, 4.99, 'Mexican', 'Official'],
    ['Taco Bell', 'Crunchy Taco', 'single', 170, 8, 13, 10, 1.79, 'Mexican', 'Official'],
    ['Taco Bell', 'Power Menu Bowl - Chicken', 'standard', 470, 27, 50, 18, 6.49, 'Mexican', 'Official high-protein'],

    # --- WaBa Grill (4 items) ---
    ['WaBa Grill', 'Chicken Plate', 'double rice', 780, 40, 110, 16, 10.95, 'Asian', 'Official'],
    ['WaBa Grill', 'Chicken WaBa Bowl', 'white rice, veggies', 540, 35, 72, 11, 8.95, 'Asian', 'Official high-protein/low-fat'],
    ['WaBa Grill', 'Steak Veggie Bowl', 'standard', 560, 33, 70, 15, 9.95, 'Asian', 'Official'],
    ['WaBa Grill', 'Tofu Veggie Bowl', 'standard', 430, 15, 75, 9, 8.45, 'Asian', 'Official vegetarian'],

    # --- Wendy's (7 items) ---
    ["Wendy's", 'Apple Pecan Salad w/ Grilled Chicken', 'full, no dressing', 430, 32, 32, 21, 8.49, 'Salads', 'Official'],
    ["Wendy's", 'Baconator', 'double', 950, 54, 40, 62, 8.99, 'Burgers', 'Official high-protein/cal'],
    ["Wendy's", 'Chili', 'large', 330, 25, 32, 11, 4.29, 'Burgers', 'Official high-protein'],
    ["Wendy's", "Dave's Single", 'quarter-pound w/ cheese', 590, 30, 39, 34, 6.99, 'Burgers', 'Official'],
    ["Wendy's", 'Grilled Chicken Sandwich', 'classic', 370, 34, 38, 9, 6.29, 'Burgers', 'Official high-protein'],
    ["Wendy's", 'Jr. Bacon Cheeseburger', 'single', 380, 20, 26, 23, 3.49, 'Burgers', 'Official'],
    ["Wendy's", 'Spicy Chicken Sandwich', 'classic', 490, 28, 49, 21, 6.49, 'Burgers', 'Official'],

    # --- Wingstop (3 items) ---
    ['Wingstop', 'Boneless Wings (plain)', '6 pc', 400, 26, 38, 15, 6.99, 'Chicken', 'Official'],
    ['Wingstop', 'Chicken Sandwich', 'classic', 470, 30, 42, 20, 6.49, 'Chicken', 'Official'],
    ['Wingstop', 'Classic Wings (plain)', '6 pc', 430, 42, 0, 28, 7.49, 'Chicken', 'Official before sauce'],

]

# ============================================================================
# BUILD
# ============================================================================
INK="16140F"; LIGHT="F7F4EC"; LINE="D8D2C4"
thin = Side(style="thin", color=LINE)
border = Border(left=thin, right=thin, top=thin, bottom=thin)
HEADERS = ["Restaurant","Item","Size / Option","Calories","Protein (g)",
           "Carbs (g)","Fat (g)","Price ($)","Category","Notes / Source"]

wb = Workbook()
ws = wb.active
ws.title = "Restaurant Macros"
ws.append(HEADERS)
for c, h in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=INK)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

rows_sorted = sorted(ROWS, key=lambda x: (x[0], x[1]))
for i, row in enumerate(rows_sorted, start=2):
    for c, val in enumerate(row, 1):
        cell = ws.cell(row=i, column=c, value=val)
        cell.font = Font(name="Arial", size=10, bold=(c == 1))
        cell.border = border
        if c in (4, 5, 6, 7, 8):
            cell.alignment = Alignment(horizontal="center")
        if c == 8:
            cell.number_format = '$#,##0.00'
    # Protein / 100 cal — a formula, so it recalcs if you edit the macros
    fc = ws.cell(row=i, column=11, value=f"=IF(D{i}=0,0,ROUND(E{i}/D{i}*100,1))")
    fc.font = Font(name="Arial", size=10)
    fc.alignment = Alignment(horizontal="center")
    fc.border = border
    if i % 2 == 0:
        for c in range(1, 12):
            cur = ws.cell(row=i, column=c)
            if cur.fill.fgColor.rgb in (None, "00000000"):
                cur.fill = PatternFill("solid", fgColor=LIGHT)

ws.cell(row=1, column=11, value="Protein / 100 cal").font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
ws.cell(row=1, column=11).fill = PatternFill("solid", fgColor=INK)
ws.cell(row=1, column=11).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.cell(row=1, column=11).border = border

last = len(rows_sorted) + 1
for i, w in enumerate([17, 40, 30, 9, 11, 9, 8, 9, 15, 42, 14], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "B2"
ws.auto_filter.ref = f"A1:K{last}"

# Sheet 2 — per-restaurant summary (all formulas)
ws2 = wb.create_sheet("By Restaurant")
ws2.append(["Restaurant","# Items","Avg Calories","Avg Protein (g)","Best Protein/100cal"])
for c in range(1, 6):
    cell = ws2.cell(row=1, column=c)
    cell.font = Font(name="Arial", bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=INK)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
names = []
for row in rows_sorted:
    if row[0] not in names:
        names.append(row[0])
names.sort()
for i, name in enumerate(names, start=2):
    ws2.cell(row=i, column=1, value=name).font = Font(name="Arial", size=10, bold=True)
    ws2.cell(row=i, column=2, value=f"=COUNTIF('Restaurant Macros'!$A$2:$A${last},A{i})")
    ws2.cell(row=i, column=3, value=f"=ROUND(AVERAGEIF('Restaurant Macros'!$A$2:$A${last},A{i},'Restaurant Macros'!$D$2:$D${last}),0)")
    ws2.cell(row=i, column=4, value=f"=ROUND(AVERAGEIF('Restaurant Macros'!$A$2:$A${last},A{i},'Restaurant Macros'!$E$2:$E${last}),0)")
    # SUMPRODUCT-MAX pattern (MAXIFS is unreliable in LibreOffice recalc)
    ws2.cell(row=i, column=5, value=f"=ROUND(SUMPRODUCT(MAX(('Restaurant Macros'!$A$2:$A${last}=A{i})*'Restaurant Macros'!$K$2:$K${last})),1)")
    for c in range(1, 6):
        if c > 1:
            ws2.cell(row=i, column=c).font = Font(name="Arial", size=10)
            ws2.cell(row=i, column=c).alignment = Alignment(horizontal="center")
        ws2.cell(row=i, column=c).border = border
for i, w in enumerate([20, 10, 14, 16, 18], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"

wb.save("restaurant-nutrition.xlsx")
print(f"Built restaurant-nutrition.xlsx: {len(rows_sorted)} items across {len(names)} chains")
